"""Bounded XLSX parsing."""

from pathlib import Path


def load_xlsx(
    path: Path,
    *,
    maximum_rows: int,
    maximum_columns: int,
    maximum_nonempty_cells: int,
) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    parts: list[str] = []
    nonempty = 0
    try:
        for sheet in workbook.worksheets:
            parts.append(f"# Sheet: {sheet.title}")
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_number > maximum_rows:
                    raise ValueError("XLSX exceeds the configured row limit")
                if len(row) > maximum_columns:
                    raise ValueError("XLSX exceeds the configured column limit")
                values = ["" if value is None else str(value) for value in row]
                nonempty += sum(bool(value.strip()) for value in values)
                if nonempty > maximum_nonempty_cells:
                    raise ValueError("XLSX exceeds the configured nonempty-cell limit")
                if any(value.strip() for value in values):
                    parts.append(" | ".join(values))
    finally:
        workbook.close()
    return "\n".join(parts)
